import os
import sqlite3
import secrets
import io
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (Flask, render_template, redirect, url_for, request,
                   flash, session, g, send_file, abort)
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-prod')

AVATAR_FOLDER = os.path.join(app.root_path, 'static', 'avatars')
ALLOWED_AVATAR_EXT = {'jpg', 'jpeg', 'png'}
# Ограничение на размер аватара убрано — принимаем файлы любого размера

UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads')
ALLOWED_PHOTO_EXT = {'jpg', 'jpeg', 'png', 'heic'}
MAX_PHOTOS_PER_STAGE = 30  # максимум фото на один этап

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Пожалуйста, войдите в систему.'

DATABASE = 'otdelka.db'

DEFAULT_STAGES = [
    'Стяжка пола',
    'Штукатурка стен',
    'Шпатлёвка стен и откосов',
    'Грунтовка',
    'Укладка напольного покрытия',
    'Поклейка обоев / покраска стен',
    'Установка межкомнатных дверей',
    'Установка плинтусов',
    'Сантехническая разводка и установка приборов',
    'Электромонтажные работы',
    'Финишная уборка',
]

# Справочник единиц измерения (фиксированный, не редактируется)
UNITS = ['м2', 'м.пог.', 'шт.', 'м3', 'м']

# Единицы измерения по умолчанию для стандартных этапов
DEFAULT_UNIT_MAP = {
    'Стяжка пола':                                    'м2',
    'Штукатурка стен':                                'м2',
    'Шпатлёвка стен и откосов':                       'м2',
    'Грунтовка':                                      'м2',
    'Укладка напольного покрытия':                    'м2',
    'Поклейка обоев / покраска стен':                 'м2',
    'Установка межкомнатных дверей':                  'шт.',
    'Установка плинтусов':                            'м.пог.',
    'Сантехническая разводка и установка приборов':   'шт.',
    'Электромонтажные работы':                        'шт.',
    'Финишная уборка':                                'м2',
}


# ─── Database ────────────────────────────────────────────────────────────────

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        db.execute('PRAGMA journal_mode=WAL')
        db.execute('PRAGMA foreign_keys=ON')
    return db


@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()


def query_db(sql, args=(), one=False):
    cur = get_db().execute(sql, args)
    rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv


def execute_db(sql, args=()):
    db = get_db()
    cur = db.execute(sql, args)
    db.commit()
    return cur.lastrowid


def get_setting(key, default=None):
    row = query_db('SELECT value FROM settings WHERE key=?', [key], one=True)
    return row['value'] if row else default


def set_setting(key, value):
    existing = query_db('SELECT id FROM settings WHERE key=?', [key], one=True)
    if existing:
        execute_db('UPDATE settings SET value=? WHERE key=?', [str(value), key])
    else:
        execute_db('INSERT INTO settings (key,value) VALUES (?,?)', [key, str(value)])


def run_migrations(db):
    """Apply schema migrations safely on existing databases."""
    # stages_template.work_description
    cols = [r[1] for r in db.execute("PRAGMA table_info(stages_template)").fetchall()]
    if 'work_description' not in cols:
        db.execute("ALTER TABLE stages_template ADD COLUMN work_description TEXT")

    # apartment_stages.work_description
    cols = [r[1] for r in db.execute("PRAGMA table_info(apartment_stages)").fetchall()]
    if 'work_description' not in cols:
        db.execute("ALTER TABLE apartment_stages ADD COLUMN work_description TEXT")

    # users: avatar, is_approved, created_at
    cols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
    if 'avatar' not in cols:
        db.execute("ALTER TABLE users ADD COLUMN avatar TEXT")
    if 'is_approved' not in cols:
        db.execute("ALTER TABLE users ADD COLUMN is_approved INTEGER NOT NULL DEFAULT 0")
        # Существующие пользователи уже в системе — подтверждаем всех
        db.execute("UPDATE users SET is_approved=1")
    if 'created_at' not in cols:
        # SQLite не поддерживает datetime() как DEFAULT в ALTER TABLE
        db.execute("ALTER TABLE users ADD COLUMN created_at TEXT")
        db.execute("UPDATE users SET created_at=datetime('now') WHERE created_at IS NULL")

    # apartments.created_at
    cols = [r[1] for r in db.execute("PRAGMA table_info(apartments)").fetchall()]
    if 'created_at' not in cols:
        db.execute("ALTER TABLE apartments ADD COLUMN created_at TEXT")
        # Заполняем существующие строки текущим временем
        db.execute("UPDATE apartments SET created_at=datetime('now') WHERE created_at IS NULL")

    # settings table (kept for future use)
    db.execute('''CREATE TABLE IF NOT EXISTS settings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        key TEXT UNIQUE NOT NULL,
        value TEXT NOT NULL
    )''')

    # apartment_stages: добавляем photos если нет (старая миграция)
    cols = [r[1] for r in db.execute("PRAGMA table_info(apartment_stages)").fetchall()]
    if 'photos' not in cols:
        db.execute("ALTER TABLE apartment_stages ADD COLUMN photos TEXT")
        db.commit()
        cols.append('photos')

    # apartment_stages: пересоздаём таблицу для поддержки 'rework' + новые поля
    # (ALTER TABLE не может изменить CHECK constraint в SQLite)
    if 'rework_reason' not in cols:
        db.executescript('''
            PRAGMA foreign_keys=OFF;

            CREATE TABLE apartment_stages_v2 (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                apartment_id INTEGER NOT NULL REFERENCES apartments(id) ON DELETE CASCADE,
                stage_name TEXT NOT NULL,
                order_num INTEGER NOT NULL DEFAULT 0,
                volume_sqm REAL,
                deadline TEXT,
                status TEXT NOT NULL DEFAULT 'not_started'
                    CHECK(status IN ('not_started','in_progress','done','overdue','rework')),
                started_at TEXT,
                completed_at TEXT,
                work_description TEXT,
                photos TEXT,
                rework_reason TEXT,
                rework_count INTEGER NOT NULL DEFAULT 0
            );

            INSERT INTO apartment_stages_v2
                (id, apartment_id, stage_name, order_num, volume_sqm, deadline,
                 status, started_at, completed_at, work_description, photos,
                 rework_reason, rework_count)
            SELECT
                id, apartment_id, stage_name, order_num, volume_sqm, deadline,
                status, started_at, completed_at, work_description, photos,
                NULL, 0
            FROM apartment_stages;

            DROP TABLE apartment_stages;
            ALTER TABLE apartment_stages_v2 RENAME TO apartment_stages;

            PRAGMA foreign_keys=ON;
        ''')

    # apartments.completed_at
    cols = [r[1] for r in db.execute("PRAGMA table_info(apartments)").fetchall()]
    if 'completed_at' not in cols:
        db.execute("ALTER TABLE apartments ADD COLUMN completed_at TEXT")

    # stages_template.default_unit
    cols = [r[1] for r in db.execute("PRAGMA table_info(stages_template)").fetchall()]
    if 'default_unit' not in cols:
        db.execute("ALTER TABLE stages_template ADD COLUMN default_unit TEXT NOT NULL DEFAULT 'м2'")

    # apartment_stages.unit
    cols = [r[1] for r in db.execute("PRAGMA table_info(apartment_stages)").fetchall()]
    if 'unit' not in cols:
        db.execute("ALTER TABLE apartment_stages ADD COLUMN unit TEXT NOT NULL DEFAULT 'м2'")

    db.commit()

    # Создать папки для файлов
    os.makedirs(AVATAR_FOLDER, exist_ok=True)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def init_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.execute('PRAGMA foreign_keys=ON')

    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','foreman','guest')),
            full_name TEXT NOT NULL DEFAULT '',
            avatar TEXT,
            is_approved INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS buildings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            address TEXT
        );

        CREATE TABLE IF NOT EXISTS apartments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            building_id INTEGER NOT NULL REFERENCES buildings(id) ON DELETE CASCADE,
            number TEXT NOT NULL,
            plan_start_date TEXT,
            plan_end_date TEXT,
            foreman_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TEXT DEFAULT (datetime('now')),
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS stages_template (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            default_order INTEGER NOT NULL DEFAULT 0,
            work_description TEXT,
            default_unit TEXT NOT NULL DEFAULT 'м2'
        );

        CREATE TABLE IF NOT EXISTS apartment_stages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            apartment_id INTEGER NOT NULL REFERENCES apartments(id) ON DELETE CASCADE,
            stage_name TEXT NOT NULL,
            order_num INTEGER NOT NULL DEFAULT 0,
            volume_sqm REAL,
            unit TEXT NOT NULL DEFAULT 'м2',
            deadline TEXT,
            status TEXT NOT NULL DEFAULT 'not_started'
                CHECK(status IN ('not_started','in_progress','done','overdue','rework')),
            started_at TEXT,
            completed_at TEXT,
            work_description TEXT,
            photos TEXT,
            rework_reason TEXT,
            rework_count INTEGER NOT NULL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS guest_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            building_id INTEGER NOT NULL REFERENCES buildings(id) ON DELETE CASCADE,
            token TEXT UNIQUE NOT NULL,
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key TEXT UNIQUE NOT NULL,
            value TEXT NOT NULL
        );
    ''')
    db.commit()

    # Run migrations for existing databases (adds columns if missing)
    run_migrations(db)

    # Seed only if empty
    count = db.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    if count > 0:
        db.close()
        return

    # Users (все seed-пользователи сразу подтверждены)
    users = [
        ('admin',    generate_password_hash('admin123'), 'admin',   'Администратор', 1),
        ('foreman1', generate_password_hash('123456'),   'foreman', 'Иванов Иван',   1),
        ('foreman2', generate_password_hash('123456'),   'foreman', 'Петров Пётр',   1),
    ]
    db.executemany(
        'INSERT INTO users (username,password_hash,role,full_name,is_approved) VALUES (?,?,?,?,?)',
        users
    )

    # Stages template with sample descriptions
    stage_descriptions = {
        'Стяжка пола': 'Выравнивание основания, заливка стяжки М300. Толщина 50–70 мм. Выдержка 28 дней до полного набора прочности.',
        'Штукатурка стен': 'Машинная или ручная штукатурка по маякам. Слой 10–20 мм. Материал — гипсовая или цементно-песчаная смесь.',
        'Шпатлёвка стен и откосов': 'Финишная шпатлёвка в 2 слоя. Шлифовка между слоями. Обязательная обработка откосов и углов.',
        'Грунтовка': 'Грунтование всех поверхностей перед финишными работами. Глубокое проникновение. Дать высохнуть 4–6 часов.',
        'Укладка напольного покрытия': 'Укладка плитки, ламината или иного покрытия согласно проекту. Соблюдать зазоры и уровень.',
        'Поклейка обоев / покраска стен': 'Поклейка обоев встык или покраска в 2 слоя. Соблюдать направление рисунка и перекрытие швов.',
        'Установка межкомнатных дверей': 'Монтаж коробок, навешивание полотен, установка наличников и фурнитуры.',
        'Установка плинтусов': 'Монтаж напольных и потолочных плинтусов. Подрезка углов. Фиксация на клей или дюбели.',
        'Сантехническая разводка и установка приборов': 'Подключение смесителей, унитаза, ванны/душевой кабины. Проверка на протечки.',
        'Электромонтажные работы': 'Установка розеток, выключателей, светильников. Подключение щитка. Прозвонка цепей.',
        'Финишная уборка': 'Влажная уборка всех поверхностей. Удаление строительного мусора. Мытьё окон.',
    }
    for i, name in enumerate(DEFAULT_STAGES):
        db.execute(
            'INSERT INTO stages_template (name,default_order,work_description,default_unit) VALUES (?,?,?,?)',
            (name, i + 1, stage_descriptions.get(name, ''), DEFAULT_UNIT_MAP.get(name, 'м2'))
        )

    # Building
    db.execute("INSERT INTO buildings (name,address) VALUES ('Солнечный','ул. Солнечная, 1')")
    building_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

    foreman1_id = db.execute("SELECT id FROM users WHERE username='foreman1'").fetchone()[0]
    foreman2_id = db.execute("SELECT id FROM users WHERE username='foreman2'").fetchone()[0]

    today = date.today()
    stage_days = 3  # default
    apts = [
        (building_id, '1', str(today), str(today + timedelta(days=60)), foreman1_id),
        (building_id, '2', str(today + timedelta(days=5)), str(today + timedelta(days=75)), foreman1_id),
        (building_id, '3', str(today + timedelta(days=10)), str(today + timedelta(days=80)), foreman2_id),
    ]
    for apt_row in apts:
        db.execute(
            'INSERT INTO apartments (building_id,number,plan_start_date,plan_end_date,foreman_id) VALUES (?,?,?,?,?)',
            apt_row
        )
        apt_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]
        plan_start_str = apt_row[2]
        plan_start_dt = date.fromisoformat(plan_start_str) if plan_start_str else None

        for i, stage_name in enumerate(DEFAULT_STAGES):
            order = i + 1
            if plan_start_dt:
                deadline = str(plan_start_dt + timedelta(days=order * stage_days))
            else:
                deadline = None
            desc = stage_descriptions.get(stage_name, '')
            db.execute(
                '''INSERT INTO apartment_stages
                   (apartment_id,stage_name,order_num,volume_sqm,deadline,status,work_description)
                   VALUES (?,?,?,?,?,?,?)''',
                (apt_id, stage_name, order, 30.0, deadline, 'not_started', desc)
            )

    db.commit()
    db.close()


# ─── User model ──────────────────────────────────────────────────────────────

class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.username = row['username']
        self.role = row['role']
        self.full_name = row['full_name']
        self.avatar      = row['avatar']      if 'avatar'      in row.keys() else None
        self.is_approved = row['is_approved'] if 'is_approved' in row.keys() else 1

    @property
    def initials(self):
        name = self.full_name or self.username
        parts = name.strip().split()
        if len(parts) >= 2:
            return (parts[0][0] + parts[1][0]).upper()
        return name[:2].upper() if len(name) >= 2 else name[:1].upper()


@app.context_processor
def inject_pending_count():
    """Передаёт pending_count во все шаблоны (для бейджа в меню)."""
    try:
        if current_user.is_authenticated and current_user.role == 'admin':
            c = query_db('SELECT COUNT(*) as c FROM users WHERE is_approved=0', one=True)
            return {'pending_count': c['c'] if c else 0}
    except Exception:
        pass
    return {'pending_count': 0}


@login_manager.user_loader
def load_user(user_id):
    row = query_db('SELECT * FROM users WHERE id=?', [user_id], one=True)
    return User(row) if row else None


# ─── Helpers ─────────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


def foreman_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('admin', 'foreman'):
            abort(403)
        return f(*args, **kwargs)
    return decorated


def save_stage_photos(stage_id, apt_id, files):
    """Сохранить фото этапа, удалив старые. Вернуть строку имён через запятую."""
    import shutil
    stage_dir = os.path.join(UPLOAD_FOLDER, str(apt_id), str(stage_id))
    if os.path.exists(stage_dir):
        shutil.rmtree(stage_dir)
    os.makedirs(stage_dir, exist_ok=True)
    filenames = []
    for i, f in enumerate(files, 1):
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'jpg'
        filename = f'photo_{i}.{ext}'
        f.save(os.path.join(stage_dir, filename))
        filenames.append(filename)
    return ','.join(filenames)


def validate_photos(files):
    """Проверить список файлов. Вернуть (ok: bool, error: str | None)."""
    valid = [f for f in files if f and f.filename]
    if not valid:
        return False, 'Прикрепите минимум 1 фотографию.'
    if len(valid) > MAX_PHOTOS_PER_STAGE:
        return False, f'Максимум {MAX_PHOTOS_PER_STAGE} фотографий.'
    for f in valid:
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_PHOTO_EXT:
            return False, f'Недопустимый формат: {f.filename}. Допустимо: JPG, PNG, HEIC.'
    return True, None


def check_apartment_completion(apt_id):
    """Обновить completed_at квартиры: заполнить если все этапы done, очистить иначе."""
    stages = query_db(
        'SELECT status FROM apartment_stages WHERE apartment_id=?', [apt_id]
    )
    if stages and all(s['status'] == 'done' for s in stages):
        execute_db(
            "UPDATE apartments SET completed_at=datetime('now') WHERE id=? AND completed_at IS NULL",
            [apt_id]
        )
    else:
        execute_db("UPDATE apartments SET completed_at=NULL WHERE id=?", [apt_id])


def refresh_overdue():
    """Mark stages as overdue only when deadline is set and passed and not done."""
    today = str(date.today())
    execute_db(
        """UPDATE apartment_stages SET status='overdue'
           WHERE status='not_started'
             AND deadline IS NOT NULL AND deadline != ''
             AND deadline < ?""",
        [today]
    )
    execute_db(
        """UPDATE apartment_stages SET status='overdue'
           WHERE status='in_progress'
             AND deadline IS NOT NULL AND deadline != ''
             AND deadline < ?""",
        [today]
    )


STATUS_LABEL = {
    'not_started': ('Не начат',     'secondary'),
    'in_progress': ('В работе',     'warning'),
    'done':        ('Завершён',     'success'),
    'overdue':     ('Просрочен',    'danger'),
    'rework':      ('На доработку', 'warning'),   # текст-dark добавляется в шаблоне
}

app.jinja_env.globals['STATUS_LABEL'] = STATUS_LABEL
app.jinja_env.globals['now'] = datetime.now
app.jinja_env.globals['UNITS'] = UNITS


# ─── Auth routes ─────────────────────────────────────────────────────────────

@app.route('/', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return _redirect_by_role(current_user.role)

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        row = query_db('SELECT * FROM users WHERE username=?', [username], one=True)
        if row and check_password_hash(row['password_hash'], password):
            if not row['is_approved']:
                flash('Ваша учётная запись ещё не подтверждена администратором. '
                      'Пожалуйста, ожидайте.', 'warning')
                return render_template('login.html')
            user = User(row)
            login_user(user)
            return _redirect_by_role(user.role)
        flash('Неверный логин или пароль.', 'danger')

    return render_template('login.html')


def _redirect_by_role(role):
    if role == 'admin':
        return redirect(url_for('admin_dashboard'))
    if role == 'foreman':
        return redirect(url_for('foreman_dashboard'))
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return _redirect_by_role(current_user.role)

    error = {}
    form  = {}

    if request.method == 'POST':
        form['username']  = request.form.get('username', '').strip()
        form['password']  = request.form.get('password', '')
        form['password2'] = request.form.get('password2', '')
        form['role']      = request.form.get('role', 'foreman')
        form['full_name'] = request.form.get('full_name', '').strip()

        if len(form['username']) < 3:
            error['username'] = 'Логин должен содержать минимум 3 символа.'
        elif query_db('SELECT id FROM users WHERE username=?', [form['username']], one=True):
            error['username'] = 'Пользователь с таким логином уже существует.'

        if len(form['password']) < 4:
            error['password'] = 'Пароль должен содержать минимум 4 символа.'
        elif form['password'] != form['password2']:
            error['password2'] = 'Пароли не совпадают.'

        if form['role'] not in ('foreman', 'guest'):
            error['role'] = 'Недопустимая роль.'

        if not error:
            execute_db(
                '''INSERT INTO users
                   (username, password_hash, role, full_name, is_approved, created_at)
                   VALUES (?,?,?,?,0,datetime('now'))''',
                [form['username'], generate_password_hash(form['password']),
                 form['role'], form['full_name']]
            )
            flash('Регистрация прошла успешно. Ожидайте подтверждения администратора. '
                  'После подтверждения вы сможете войти.', 'success')
            return redirect(url_for('login'))

    return render_template('register.html', error=error, form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


# ─── Guest token route ────────────────────────────────────────────────────────

@app.route('/guest/<token>')
def guest_view(token):
    gt = query_db('SELECT * FROM guest_tokens WHERE token=?', [token], one=True)
    if not gt:
        abort(404)
    refresh_overdue()
    building = query_db('SELECT * FROM buildings WHERE id=?', [gt['building_id']], one=True)
    apartments = query_db(
        'SELECT * FROM apartments WHERE building_id=? ORDER BY number', [gt['building_id']]
    )

    # Собираем данные по всем квартирам
    all_apt_data = []
    for apt in apartments:
        stages = query_db(
            'SELECT * FROM apartment_stages WHERE apartment_id=? ORDER BY order_num',
            [apt['id']]
        )
        total = len(stages)
        done  = sum(1 for s in stages if s['status'] == 'done')
        pct   = int(done / total * 100) if total else 0
        has_active  = any(s['status'] in ('in_progress', 'rework') for s in stages)
        has_overdue = any(s['status'] == 'overdue' for s in stages)
        all_apt_data.append({
            'apt': apt, 'stages': stages, 'pct': pct,
            'has_active': has_active, 'has_overdue': has_overdue,
        })

    # Статистика по всем квартирам (не меняется при фильтре)
    stats = {
        'total':    len(all_apt_data),
        'active':   sum(1 for d in all_apt_data if d['has_active'] and not d['apt']['completed_at']),
        'done':     sum(1 for d in all_apt_data if d['apt']['completed_at']),
        'overdue':  sum(1 for d in all_apt_data if d['has_overdue']),
    }

    # Применяем фильтр
    f = request.args.get('filter', 'all')
    if f == 'active':
        apt_data = [d for d in all_apt_data if d['has_active'] and not d['apt']['completed_at']]
    elif f == 'completed':
        apt_data = [d for d in all_apt_data if d['apt']['completed_at']]
    elif f == 'overdue':
        apt_data = [d for d in all_apt_data if d['has_overdue']]
    else:
        f = 'all'
        apt_data = all_apt_data

    return render_template('guest/view.html',
                           building=building, apt_data=apt_data,
                           stats=stats, active_filter=f, token=token)


@app.route('/guest/<token>/export')
def guest_export(token):
    """Скачать Excel-сводку по ЖК для гостя."""
    gt = query_db('SELECT * FROM guest_tokens WHERE token=?', [token], one=True)
    if not gt:
        abort(404)
    building = query_db('SELECT * FROM buildings WHERE id=?', [gt['building_id']], one=True)

    rows = query_db(
        '''SELECT a.number, s.stage_name, s.status, s.volume_sqm, s.unit,
                  s.deadline, s.completed_at
           FROM apartment_stages s
           JOIN apartments a ON s.apartment_id = a.id
           WHERE a.building_id=?
           ORDER BY a.number, s.order_num''',
        [gt['building_id']]
    )

    STATUS_RU = {
        'not_started': 'Не начат',
        'in_progress': 'В работе',
        'done':        'Завершён',
        'overdue':     'Просрочен',
        'rework':      'На доработку',
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Сводка'

    header_fill = PatternFill(fill_type='solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF')

    headers = ['Квартира', 'Этап', 'Статус', 'Объём', 'Ед. изм.', 'Дедлайн', 'Завершён']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row['number'])
        ws.cell(row=i, column=2, value=row['stage_name'])
        ws.cell(row=i, column=3, value=STATUS_RU.get(row['status'], row['status']))
        ws.cell(row=i, column=4, value=row['volume_sqm'])
        ws.cell(row=i, column=5, value=row['unit'] or 'м2')
        ws.cell(row=i, column=6, value=row['deadline'] or '')
        completed = row['completed_at']
        ws.cell(row=i, column=7, value=completed[:10] if completed else '')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'сводка_{building["name"]}.xlsx'.replace(' ', '_')
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─── Admin: Dashboard ────────────────────────────────────────────────────────

@app.route('/admin/')
@login_required
@admin_required
def admin_dashboard():
    refresh_overdue()

    # Считаем уникальные квартиры по комбинации (building_id, number)
    total = query_db(
        "SELECT COUNT(DISTINCT building_id || '_' || number) as c FROM apartments",
        one=True
    )['c']

    in_progress = query_db(
        """SELECT COUNT(DISTINCT a.building_id || '_' || a.number) as c
           FROM apartment_stages s
           JOIN apartments a ON s.apartment_id = a.id
           WHERE s.status = 'in_progress'""",
        one=True
    )['c']

    done_apts = query_db(
        """SELECT COUNT(DISTINCT building_id || '_' || number) as c
           FROM apartments WHERE completed_at IS NOT NULL""",
        one=True
    )['c']

    overdue = query_db(
        """SELECT COUNT(DISTINCT a.building_id || '_' || a.number) as c
           FROM apartment_stages s
           JOIN apartments a ON s.apartment_id = a.id
           WHERE s.status = 'overdue'""",
        one=True
    )['c']

    buildings = query_db('SELECT * FROM buildings ORDER BY name')
    return render_template('admin/dashboard.html',
                           total=total, in_progress=in_progress,
                           done_apts=done_apts, overdue=overdue,
                           buildings=buildings)


# ─── Admin: Buildings ────────────────────────────────────────────────────────

@app.route('/admin/buildings')
@login_required
@admin_required
def admin_buildings():
    buildings = query_db('SELECT * FROM buildings ORDER BY name')
    return render_template('admin/buildings.html', buildings=buildings)


@app.route('/admin/buildings/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_building_add():
    if request.method == 'POST':
        name = request.form['name'].strip()
        address = request.form.get('address', '').strip()
        if name:
            execute_db('INSERT INTO buildings (name,address) VALUES (?,?)', [name, address])
            flash('ЖК создан.', 'success')
            return redirect(url_for('admin_buildings'))
    return render_template('admin/building_form.html', building=None)


@app.route('/admin/buildings/<int:bid>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_building_edit(bid):
    building = query_db('SELECT * FROM buildings WHERE id=?', [bid], one=True)
    if not building:
        abort(404)
    if request.method == 'POST':
        name = request.form['name'].strip()
        address = request.form.get('address', '').strip()
        execute_db('UPDATE buildings SET name=?,address=? WHERE id=?', [name, address, bid])
        flash('ЖК обновлён.', 'success')
        return redirect(url_for('admin_buildings'))
    return render_template('admin/building_form.html', building=building)


@app.route('/admin/buildings/<int:bid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_building_delete(bid):
    execute_db('DELETE FROM buildings WHERE id=?', [bid])
    flash('ЖК удалён.', 'success')
    return redirect(url_for('admin_buildings'))


# ─── Admin: Apartments ───────────────────────────────────────────────────────

@app.route('/admin/buildings/<int:bid>/apartments')
@login_required
@admin_required
def admin_apartments(bid):
    building = query_db('SELECT * FROM buildings WHERE id=?', [bid], one=True)
    if not building:
        abort(404)
    apartments = query_db(
        '''SELECT a.*, u.full_name as foreman_name
           FROM apartments a LEFT JOIN users u ON a.foreman_id=u.id
           WHERE a.building_id=? ORDER BY a.number''',
        [bid]
    )
    apt_stats = {}
    for apt in apartments:
        stages = query_db(
            'SELECT status FROM apartment_stages WHERE apartment_id=?', [apt['id']]
        )
        total = len(stages)
        done = sum(1 for s in stages if s['status'] == 'done')
        apt_stats[apt['id']] = {'total': total, 'done': done,
                                 'pct': int(done / total * 100) if total else 0}

    # Находим дублирующиеся номера квартир в этом ЖК
    dup_rows = query_db(
        '''SELECT number FROM apartments WHERE building_id=?
           GROUP BY number HAVING COUNT(*) > 1''',
        [bid]
    )
    duplicate_numbers = {r['number'] for r in dup_rows}

    return render_template('admin/apartments.html',
                           building=building, apartments=apartments,
                           apt_stats=apt_stats, duplicate_numbers=duplicate_numbers)


@app.route('/admin/buildings/<int:bid>/apartments/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_apartment_add(bid):
    building = query_db('SELECT * FROM buildings WHERE id=?', [bid], one=True)
    foremen  = query_db("SELECT * FROM users WHERE role='foreman' ORDER BY full_name")

    if request.method == 'POST':
        number     = request.form['number'].strip()
        plan_start = request.form.get('plan_start_date') or None
        plan_end   = request.form.get('plan_end_date') or None
        foreman_id = request.form.get('foreman_id') or None

        apt_id = execute_db(
            '''INSERT INTO apartments
               (building_id, number, plan_start_date, plan_end_date, foreman_id, created_at)
               VALUES (?,?,?,?,?,datetime('now'))''',
            [bid, number, plan_start, plan_end, foreman_id]
        )

        templates = query_db('SELECT * FROM stages_template ORDER BY default_order')
        added = 0
        for t in templates:
            if not request.form.get(f'stage_sel_{t["id"]}'):
                continue  # чекбокс не отмечен

            # Дедлайн — только из поля формы (NULL если не заполнено)
            deadline_raw = request.form.get(f'stage_deadline_{t["id"]}', '').strip()
            deadline = deadline_raw if deadline_raw else None

            volume_raw = request.form.get(f'stage_volume_{t["id"]}', '').strip()
            try:
                volume = float(volume_raw) if volume_raw else None
            except ValueError:
                volume = None

            unit = request.form.get(f'stage_unit_{t["id"]}', '').strip()
            if unit not in UNITS:
                unit = t['default_unit'] if t['default_unit'] else 'м2'

            execute_db(
                '''INSERT INTO apartment_stages
                   (apartment_id, stage_name, order_num, volume_sqm, unit, deadline, status, work_description)
                   VALUES (?,?,?,?,?,?,?,?)''',
                [apt_id, t['name'], t['default_order'], volume, unit, deadline,
                 'not_started', t['work_description'] or '']
            )
            added += 1

        flash(f'Квартира создана, добавлено {added} этапов.', 'success')
        return redirect(url_for('admin_apartments', bid=bid))

    templates = query_db('SELECT * FROM stages_template ORDER BY default_order')
    return render_template('admin/apartment_form.html',
                           building=building, apartment=None, foremen=foremen,
                           templates=templates)


@app.route('/admin/apartments/<int:apt_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_apartment_edit(apt_id):
    apt = query_db('SELECT * FROM apartments WHERE id=?', [apt_id], one=True)
    if not apt:
        abort(404)
    building = query_db('SELECT * FROM buildings WHERE id=?', [apt['building_id']], one=True)
    foremen = query_db("SELECT * FROM users WHERE role='foreman' ORDER BY full_name")
    if request.method == 'POST':
        number = request.form['number'].strip()
        plan_start = request.form.get('plan_start_date') or None
        plan_end = request.form.get('plan_end_date') or None
        foreman_id = request.form.get('foreman_id') or None
        execute_db(
            'UPDATE apartments SET number=?,plan_start_date=?,plan_end_date=?,foreman_id=? WHERE id=?',
            [number, plan_start, plan_end, foreman_id, apt_id]
        )
        flash('Квартира обновлена. Дедлайны этапов не изменены.', 'success')
        return redirect(url_for('admin_apartments', bid=apt['building_id']))
    return render_template('admin/apartment_form.html',
                           building=building, apartment=apt, foremen=foremen)


@app.route('/admin/apartments/<int:apt_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_apartment_delete(apt_id):
    apt = query_db('SELECT * FROM apartments WHERE id=?', [apt_id], one=True)
    if not apt:
        abort(404)
    bid = apt['building_id']
    execute_db('DELETE FROM apartments WHERE id=?', [apt_id])
    flash('Квартира удалена.', 'success')
    return redirect(url_for('admin_apartments', bid=bid))


# ─── Admin: Stages of apartment ──────────────────────────────────────────────

@app.route('/admin/apartments/<int:apt_id>/stages')
@login_required
@admin_required
def admin_stages(apt_id):
    refresh_overdue()
    apt = query_db('SELECT * FROM apartments WHERE id=?', [apt_id], one=True)
    if not apt:
        abort(404)
    building = query_db('SELECT * FROM buildings WHERE id=?', [apt['building_id']], one=True)
    stages = query_db(
        'SELECT * FROM apartment_stages WHERE apartment_id=? ORDER BY order_num', [apt_id]
    )
    # Этапы справочника, которых ещё нет в квартире — для модалки добавления
    existing_names = {s['stage_name'] for s in stages}
    available = query_db('SELECT * FROM stages_template ORDER BY default_order')
    available = [t for t in available if t['name'] not in existing_names]
    return render_template('admin/stages.html', apt=apt, building=building,
                           stages=stages, available=available)


@app.route('/admin/apartments/<int:apt_id>/stages/add', methods=['POST'])
@login_required
@admin_required
def admin_stage_add(apt_id):
    """Добавить один этап в существующую квартиру (через модальное окно)."""
    apt = query_db('SELECT * FROM apartments WHERE id=?', [apt_id], one=True)
    if not apt:
        abort(404)
    stage_name = request.form.get('stage_name', '').strip()
    deadline   = request.form.get('deadline') or None
    volume     = request.form.get('volume_sqm') or None
    if volume:
        try:
            volume = float(volume)
        except ValueError:
            volume = None

    unit = request.form.get('unit', '').strip()
    if unit not in UNITS:
        unit = 'м2'

    if stage_name:
        max_ord = query_db(
            'SELECT COALESCE(MAX(order_num), 0) as m FROM apartment_stages WHERE apartment_id=?',
            [apt_id], one=True
        )['m']
        # Берём work_description и default_unit из справочника, если есть
        tpl = query_db('SELECT work_description, default_unit FROM stages_template WHERE name=?',
                       [stage_name], one=True)
        work_desc = tpl['work_description'] if tpl and tpl['work_description'] else ''
        if unit == 'м2' and tpl and tpl['default_unit']:
            unit = tpl['default_unit']
        execute_db(
            '''INSERT INTO apartment_stages
               (apartment_id, stage_name, order_num, volume_sqm, unit, deadline, status, work_description)
               VALUES (?,?,?,?,?,?,?,?)''',
            [apt_id, stage_name, max_ord + 1, volume, unit, deadline, 'not_started', work_desc]
        )
        # Новый этап — квартира больше не завершена
        check_apartment_completion(apt_id)
        flash(f'Этап «{stage_name}» добавлен в квартиру.', 'success')
    return redirect(url_for('admin_stages', apt_id=apt_id))


@app.route('/admin/stages/<int:stage_id>/delete', methods=['POST'])
@login_required
@admin_required
def admin_stage_delete(stage_id):
    """Удалить этап из квартиры."""
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage:
        abort(404)
    apt_id = stage['apartment_id']
    execute_db('DELETE FROM apartment_stages WHERE id=?', [stage_id])
    check_apartment_completion(apt_id)
    flash(f'Этап «{stage["stage_name"]}» удалён из квартиры.', 'success')
    return redirect(url_for('admin_stages', apt_id=apt_id))


@app.route('/admin/stages/<int:stage_id>/move/<direction>', methods=['POST'])
@login_required
@admin_required
def admin_stage_move(stage_id, direction):
    """Переместить этап вверх (up) или вниз (down), обменяв order_num с соседом."""
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage or direction not in ('up', 'down'):
        abort(404)
    apt_id = stage['apartment_id']
    cur_order = stage['order_num']

    if direction == 'up':
        neighbor = query_db(
            '''SELECT * FROM apartment_stages
               WHERE apartment_id=? AND order_num < ?
               ORDER BY order_num DESC LIMIT 1''',
            [apt_id, cur_order], one=True
        )
    else:
        neighbor = query_db(
            '''SELECT * FROM apartment_stages
               WHERE apartment_id=? AND order_num > ?
               ORDER BY order_num ASC LIMIT 1''',
            [apt_id, cur_order], one=True
        )

    if neighbor:
        execute_db('UPDATE apartment_stages SET order_num=? WHERE id=?',
                   [neighbor['order_num'], stage_id])
        execute_db('UPDATE apartment_stages SET order_num=? WHERE id=?',
                   [cur_order, neighbor['id']])
    return redirect(url_for('admin_stages', apt_id=apt_id))


@app.route('/admin/stages/<int:stage_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_stage_edit(stage_id):
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage:
        abort(404)
    apt = query_db('SELECT * FROM apartments WHERE id=?', [stage['apartment_id']], one=True)
    building = query_db('SELECT * FROM buildings WHERE id=?', [apt['building_id']], one=True)
    if request.method == 'POST':
        volume = request.form.get('volume_sqm') or None
        deadline = request.form.get('deadline') or None
        work_desc = request.form.get('work_description', '').strip()
        unit = request.form.get('unit', '').strip()
        if unit not in UNITS:
            unit = 'м2'
        if volume:
            volume = float(volume)
        execute_db(
            'UPDATE apartment_stages SET volume_sqm=?,unit=?,deadline=?,work_description=? WHERE id=?',
            [volume, unit, deadline, work_desc, stage_id]
        )
        flash('Этап обновлён.', 'success')
        return redirect(url_for('admin_stages', apt_id=stage['apartment_id']))
    return render_template('admin/stage_edit.html', stage=stage, apt=apt, building=building)


# ─── Admin: Stages template ──────────────────────────────────────────────────

@app.route('/admin/stages-template')
@login_required
@admin_required
def admin_stages_template():
    templates = query_db('SELECT * FROM stages_template ORDER BY default_order')
    return render_template('admin/stages_template.html', templates=templates)


@app.route('/admin/stages-template/add', methods=['POST'])
@login_required
@admin_required
def admin_stages_template_add():
    name = request.form.get('name', '').strip()
    default_unit = request.form.get('default_unit', 'м2').strip()
    if default_unit not in UNITS:
        default_unit = 'м2'
    if name:
        max_order = query_db('SELECT MAX(default_order) as m FROM stages_template', one=True)['m'] or 0
        execute_db('INSERT INTO stages_template (name,default_order,work_description,default_unit) VALUES (?,?,?,?)',
                   [name, max_order + 1, '', default_unit])
        flash('Этап добавлен в справочник.', 'success')
    return redirect(url_for('admin_stages_template'))


@app.route('/admin/stages-template/<int:tid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_stages_template_delete(tid):
    execute_db('DELETE FROM stages_template WHERE id=?', [tid])
    flash('Этап удалён из справочника.', 'success')
    return redirect(url_for('admin_stages_template'))


@app.route('/admin/stages-template/<int:tid>/update', methods=['POST'])
@login_required
@admin_required
def admin_stages_template_update(tid):
    name = request.form.get('name', '').strip()
    work_desc = request.form.get('work_description', '').strip()
    default_unit = request.form.get('default_unit', 'м2').strip()
    if default_unit not in UNITS:
        default_unit = 'м2'
    if name:
        execute_db(
            'UPDATE stages_template SET name=?,work_description=?,default_unit=? WHERE id=?',
            [name, work_desc, default_unit, tid]
        )
        flash('Этап обновлён.', 'success')
    return redirect(url_for('admin_stages_template'))


# ─── Profile (все роли) ───────────────────────────────────────────────────────

@app.route('/profile')
@login_required
def profile():
    return render_template('profile.html')


@app.route('/profile/avatar', methods=['POST'])
@login_required
def profile_avatar():
    f = request.files.get('avatar')
    if not f or f.filename == '':
        flash('Выберите файл.', 'warning')
        return redirect(url_for('profile'))

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ALLOWED_AVATAR_EXT:
        flash('Допустимые форматы: JPG, JPEG, PNG.', 'danger')
        return redirect(url_for('profile'))

    data = f.read()
    os.makedirs(AVATAR_FOLDER, exist_ok=True)

    # Удаляем старый аватар (любое расширение)
    for old_ext in ALLOWED_AVATAR_EXT:
        old_path = os.path.join(AVATAR_FOLDER, f'user_{current_user.id}.{old_ext}')
        if os.path.exists(old_path):
            os.remove(old_path)

    # Сохраняем новый
    filename = f'user_{current_user.id}.{ext}'
    filepath = os.path.join(AVATAR_FOLDER, filename)
    with open(filepath, 'wb') as out:
        out.write(data)

    execute_db('UPDATE users SET avatar=? WHERE id=?', [filename, current_user.id])
    # Обновляем объект в сессии
    current_user.avatar = filename
    flash('Аватар обновлён.', 'success')
    return redirect(url_for('profile'))


@app.route('/profile/login', methods=['POST'])
@login_required
def profile_login():
    if current_user.role == 'guest':
        abort(403)
    new_login = request.form.get('new_username', '').strip()
    if len(new_login) < 3:
        flash('Логин должен содержать минимум 3 символа.', 'danger')
        return redirect(url_for('profile'))
    existing = query_db('SELECT id FROM users WHERE username=? AND id!=?',
                        [new_login, current_user.id], one=True)
    if existing:
        flash('Этот логин уже используется другим пользователем.', 'danger')
        return redirect(url_for('profile'))
    execute_db('UPDATE users SET username=? WHERE id=?', [new_login, current_user.id])
    current_user.username = new_login
    flash('Логин успешно изменён.', 'success')
    return redirect(url_for('profile'))


@app.route('/profile/password', methods=['POST'])
@login_required
def profile_password():
    current_pw  = request.form.get('current_password', '')
    new_pw      = request.form.get('new_password', '')
    confirm_pw  = request.form.get('confirm_password', '')

    row = query_db('SELECT password_hash FROM users WHERE id=?', [current_user.id], one=True)
    if not check_password_hash(row['password_hash'], current_pw):
        flash('Текущий пароль введён неверно.', 'danger')
        return redirect(url_for('profile'))
    if len(new_pw) < 4:
        flash('Новый пароль должен содержать минимум 4 символа.', 'danger')
        return redirect(url_for('profile'))
    if new_pw != confirm_pw:
        flash('Новый пароль и подтверждение не совпадают.', 'danger')
        return redirect(url_for('profile'))

    execute_db('UPDATE users SET password_hash=? WHERE id=?',
               [generate_password_hash(new_pw), current_user.id])
    flash('Пароль успешно изменён.', 'success')
    return redirect(url_for('profile'))


# ─── Admin: Guest tokens ──────────────────────────────────────────────────────

@app.route('/admin/guest-tokens')
@login_required
@admin_required
def admin_guest_tokens():
    tokens = query_db(
        '''SELECT gt.*, b.name as building_name
           FROM guest_tokens gt JOIN buildings b ON gt.building_id=b.id
           ORDER BY gt.created_at DESC'''
    )
    buildings = query_db('SELECT * FROM buildings ORDER BY name')
    return render_template('admin/guest_tokens.html', tokens=tokens, buildings=buildings)


@app.route('/admin/guest-tokens/create', methods=['POST'])
@login_required
@admin_required
def admin_guest_token_create():
    building_id = request.form.get('building_id')
    if building_id:
        token = secrets.token_urlsafe(24)
        execute_db(
            'INSERT INTO guest_tokens (building_id,token,created_at) VALUES (?,?,?)',
            [building_id, token, datetime.now().isoformat()]
        )
        flash(f'Ссылка создана: /guest/{token}', 'success')
    return redirect(url_for('admin_guest_tokens'))


@app.route('/admin/guest-tokens/<int:tid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_guest_token_delete(tid):
    execute_db('DELETE FROM guest_tokens WHERE id=?', [tid])
    flash('Гостевой доступ удалён.', 'success')
    return redirect(url_for('admin_guest_tokens'))


# ─── Admin: Export ────────────────────────────────────────────────────────────

@app.route('/admin/export', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_export():
    buildings = query_db('SELECT * FROM buildings ORDER BY name')
    if request.method == 'POST':
        building_id = request.form.get('building_id')
        date_from = request.form.get('date_from')
        date_to = request.form.get('date_to')

        rows = query_db(
            '''SELECT a.number, s.stage_name, s.volume_sqm, s.unit, s.completed_at
               FROM apartment_stages s
               JOIN apartments a ON s.apartment_id=a.id
               WHERE a.building_id=? AND s.status='done'
               AND s.completed_at IS NOT NULL
               AND substr(s.completed_at,1,10) >= ?
               AND substr(s.completed_at,1,10) <= ?
               ORDER BY a.number, s.order_num''',
            [building_id, date_from, date_to]
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Выполненные работы'

        header_fill = PatternFill(fill_type='solid', fgColor='4472C4')
        header_font_white = Font(bold=True, color='FFFFFF')

        headers = ['Номер квартиры', 'Этап', 'Объём', 'Ед. изм.', 'Дата завершения']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row['number'])
            ws.cell(row=row_idx, column=2, value=row['stage_name'])
            ws.cell(row=row_idx, column=3, value=row['volume_sqm'])
            ws.cell(row=row_idx, column=4, value=row['unit'] or 'м2')
            completed = row['completed_at']
            if completed:
                completed = completed[:10]
            ws.cell(row=row_idx, column=5, value=completed)

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name='выполненные_работы.xlsx',
                         as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return render_template('admin/export.html', buildings=buildings)


# ─── Admin: Users ────────────────────────────────────────────────────────────

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = query_db("SELECT * FROM users ORDER BY role, username")
    return render_template('admin/users.html', users=users)


@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_user_add():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        role = request.form['role']
        full_name = request.form.get('full_name', '').strip()
        if username and password and role:
            try:
                execute_db(
                    '''INSERT INTO users
                       (username,password_hash,role,full_name,is_approved,created_at)
                       VALUES (?,?,?,?,1,datetime('now'))''',
                    [username, generate_password_hash(password), role, full_name]
                )
                flash('Пользователь создан.', 'success')
                return redirect(url_for('admin_users'))
            except Exception:
                flash('Пользователь с таким логином уже существует.', 'danger')
    return render_template('admin/user_form.html')


@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@login_required
@admin_required
def admin_user_delete(uid):
    if uid == current_user.id:
        flash('Нельзя удалить собственную учётную запись.', 'danger')
    else:
        execute_db('DELETE FROM users WHERE id=?', [uid])
        flash('Пользователь удалён.', 'success')
    return redirect(url_for('admin_users'))


@app.route('/admin/pending')
@login_required
@admin_required
def admin_pending():
    pending = query_db(
        "SELECT * FROM users WHERE is_approved=0 ORDER BY created_at DESC"
    )
    return render_template('admin/pending.html', pending=pending)


@app.route('/admin/pending/<int:uid>/approve', methods=['POST'])
@login_required
@admin_required
def admin_pending_approve(uid):
    row = query_db('SELECT username FROM users WHERE id=? AND is_approved=0', [uid], one=True)
    if row:
        execute_db('UPDATE users SET is_approved=1 WHERE id=?', [uid])
        flash(f'Пользователь «{row["username"]}» подтверждён.', 'success')
    return redirect(url_for('admin_pending'))


@app.route('/admin/pending/<int:uid>/reject', methods=['POST'])
@login_required
@admin_required
def admin_pending_reject(uid):
    row = query_db('SELECT username FROM users WHERE id=? AND is_approved=0', [uid], one=True)
    if row:
        execute_db('DELETE FROM users WHERE id=?', [uid])
        flash(f'Пользователь «{row["username"]}» отклонён и удалён.', 'warning')
    return redirect(url_for('admin_pending'))


# ─── Foreman routes ───────────────────────────────────────────────────────────

@app.route('/foreman/')
@login_required
@foreman_required
def foreman_dashboard():
    refresh_overdue()
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))

    # Сортировка: последняя созданная — первой (новые вверху)
    apartments = query_db(
        '''SELECT a.*, b.name as building_name, b.address as building_address
           FROM apartments a JOIN buildings b ON a.building_id=b.id
           WHERE a.foreman_id=?
           ORDER BY COALESCE(a.created_at, '') DESC, a.id DESC''',
        [current_user.id]
    )

    apt_data = []
    for apt in apartments:
        stages = query_db(
            'SELECT * FROM apartment_stages WHERE apartment_id=? ORDER BY order_num',
            [apt['id']]
        )
        total = len(stages)
        done  = sum(1 for s in stages if s['status'] == 'done')
        # «В работе»: есть этапы в статусе in_progress, overdue или rework, квартира не завершена
        active = (not apt['completed_at'] and
                  any(s['status'] in ('in_progress', 'overdue', 'rework') for s in stages))
        pct   = int(done / total * 100) if total else 0
        apt_data.append({
            'apt':    apt,
            'stages': stages,
            'total':  total,
            'done':   done,
            'active': active,
            'pct':    pct,
        })

    # Сводная статистика — только по квартирам этого бригадира
    stats = {
        'total':       len(apt_data),
        # «В работе»: незавершённые квартиры с активными этапами
        'in_progress': sum(1 for d in apt_data if d['active']),
        # «Завершено»: квартиры с completed_at из БД
        'completed':   sum(1 for d in apt_data if d['apt']['completed_at']),
    }

    return render_template('foreman/dashboard.html', apt_data=apt_data, stats=stats)


@app.route('/foreman/stages/<int:stage_id>/action', methods=['POST'])
@login_required
@foreman_required
def foreman_stage_action(stage_id):
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage:
        abort(404)
    apt = query_db('SELECT * FROM apartments WHERE id=?', [stage['apartment_id']], one=True)
    if current_user.role != 'admin' and apt['foreman_id'] != current_user.id:
        abort(403)

    now = datetime.now().isoformat()
    if stage['status'] in ('not_started', 'overdue'):
        # «Начать»
        execute_db(
            "UPDATE apartment_stages SET status='in_progress', started_at=? WHERE id=?",
            [now, stage_id]
        )
    elif stage['status'] == 'rework':
        # «Возобновить» — очищаем причину, возвращаем в работу
        execute_db(
            "UPDATE apartment_stages SET status='in_progress', rework_reason=NULL WHERE id=?",
            [stage_id]
        )
    # in_progress → done — через foreman_stage_finish (с фото)
    return redirect(url_for('foreman_dashboard'))


@app.route('/foreman/stages/<int:stage_id>/finish', methods=['POST'])
@login_required
@foreman_required
def foreman_stage_finish(stage_id):
    """Завершить этап с загрузкой фотографий."""
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage:
        abort(404)
    apt = query_db('SELECT * FROM apartments WHERE id=?', [stage['apartment_id']], one=True)
    if current_user.role != 'admin' and apt['foreman_id'] != current_user.id:
        abort(403)

    files = [f for f in request.files.getlist('photos') if f and f.filename]
    ok, err = validate_photos(files)
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('foreman_dashboard'))

    photos_str = save_stage_photos(stage_id, apt['id'], files)
    now = datetime.now().isoformat()
    execute_db(
        "UPDATE apartment_stages SET status='done', completed_at=?, photos=? WHERE id=?",
        [now, photos_str, stage_id]
    )
    check_apartment_completion(apt['id'])
    flash(f'Этап «{stage["stage_name"]}» завершён, загружено фото: {len(files)}.', 'success')
    return redirect(url_for('foreman_dashboard'))


@app.route('/admin/stages/<int:stage_id>/photos', methods=['POST'])
@login_required
@admin_required
def admin_stage_photos(stage_id):
    """Обновить фотографии завершённого этапа (администратор)."""
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage:
        abort(404)
    apt = query_db('SELECT * FROM apartments WHERE id=?', [stage['apartment_id']], one=True)

    files = [f for f in request.files.getlist('photos') if f and f.filename]
    ok, err = validate_photos(files)
    if not ok:
        flash(err, 'danger')
        return redirect(url_for('admin_stages', apt_id=stage['apartment_id']))

    photos_str = save_stage_photos(stage_id, apt['id'], files)
    execute_db('UPDATE apartment_stages SET photos=? WHERE id=?', [photos_str, stage_id])
    flash(f'Фотографии обновлены ({len(files)} шт.).', 'success')
    return redirect(url_for('admin_stages', apt_id=stage['apartment_id']))


@app.route('/admin/stages/<int:stage_id>/rework', methods=['POST'])
@login_required
@admin_required
def admin_stage_rework(stage_id):
    """Вернуть завершённый этап на доработку с указанием причины."""
    stage = query_db('SELECT * FROM apartment_stages WHERE id=?', [stage_id], one=True)
    if not stage or stage['status'] != 'done':
        flash('Возврат на доработку доступен только для завершённых этапов.', 'danger')
        return redirect(url_for('admin_stages', apt_id=stage['apartment_id']) if stage else url_for('admin_dashboard'))

    reason = request.form.get('rework_reason', '').strip()
    if len(reason) < 10:
        flash('Укажите причину возврата (минимум 10 символов).', 'danger')
        return redirect(url_for('admin_stages', apt_id=stage['apartment_id']))

    rework_count = (stage['rework_count'] or 0) + 1
    execute_db(
        '''UPDATE apartment_stages
           SET status='rework', completed_at=NULL, rework_reason=?, rework_count=?
           WHERE id=?''',
        [reason, rework_count, stage_id]
    )
    # Квартира больше не завершена
    check_apartment_completion(stage['apartment_id'])
    flash(f'Этап «{stage["stage_name"]}» возвращён на доработку.', 'warning')
    return redirect(url_for('admin_stages', apt_id=stage['apartment_id']))


# ─── Error handlers ───────────────────────────────────────────────────────────

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404


# ─── Entry point ─────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=8080, use_reloader=False)
